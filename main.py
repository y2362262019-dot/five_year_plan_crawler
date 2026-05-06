"""
中国地级市五年规划文件爬虫 — 多线程版

Usage:
    python -m five_year_plan_crawler.main                     # 全量爬取
    python -m five_year_plan_crawler.main --city 北京市        # 指定城市
    python -m five_year_plan_crawler.main --workers 5          # 并发数
    python -m five_year_plan_crawler.main --resume             # 断点续传
    python -m five_year_plan_crawler.main --limit 10           # 测试用
"""

import argparse
import os
import signal
import sys
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from urllib.parse import urlparse

import pandas as pd
import openpyxl

from . import config, checkpoint, extractor, organizer, rate_limiter
from .downloader import download_file
from .excel_writer import EXCEL_COLUMNS
from .fetcher import fetch_url

running = True
lock = threading.Lock()
results_lock = threading.Lock()


def _handle_signal(signum, frame):
    global running
    running = False
    print("\n收到中断信号，正在等待当前任务完成...")


signal.signal(signal.SIGINT, _handle_signal)
signal.signal(signal.SIGTERM, _handle_signal)


def _is_skip_domain(url_str: str) -> bool:
    if pd.isna(url_str):
        return True
    netloc = urlparse(str(url_str)).netloc.lower()
    for skip in config.SKIP_DOMAINS:
        if skip in netloc:
            return True
    return False


def _url_file_type(url_str: str) -> str:
    url = str(url_str).lower().rstrip("/")
    if url.endswith(".pdf"):
        return "pdf"
    if url.endswith(".doc") or url.endswith(".docx") or url.endswith(".wps"):
        return "doc"
    return "html"


def _process_one(row, limiter) -> dict:
    """Process a single row. Thread-safe."""
    if not running:
        return {"链接序号": row["链接序号"], "抓取状态": "ABORTED", "本地路径": ""}

    seq_id = row["链接序号"]
    province = str(row["省份"])
    city = str(row["城市"])
    period = str(row["所属时期"])
    url = row["新网址"]

    result = {
        "链接序号": seq_id,
        "抓取状态": "",
        "本地路径": "",
    }

    if pd.isna(url) or str(url).strip() == "":
        result["抓取状态"] = "NO_URL"
        return result

    url_str = str(url).strip()

    if _is_skip_domain(url_str):
        result["抓取状态"] = "SKIPPED_DOMAIN"
        return result

    ftype = _url_file_type(url_str)

    if ftype in ("pdf", "doc"):
        limiter.wait()
        ext = os.path.splitext(urlparse(url_str).path)[1] or ".pdf"
        fname = organizer.make_download_name(seq_id, province, city, period, ext)
        dl_result = download_file(url_str, organizer.files_subdir(province, city), fname)
        if dl_result.status == "SUCCESS":
            limiter.report_success()
        else:
            limiter.report_error()
        result["本地路径"] = dl_result.local_path
        result["抓取状态"] = dl_result.status
        return result

    # HTML page
    limiter.wait()
    fetch_result = fetch_url(url_str)
    if fetch_result.fetch_status == "SKIPPED_DOMAIN":
        result["抓取状态"] = "SKIPPED"
        return result

    if fetch_result.fetch_status != "SUCCESS":
        limiter.report_error()
        result["抓取状态"] = "FAILED"
        return result

    limiter.report_success()
    extracted = extractor.extract_page_content(fetch_result.text)
    page_title = extracted["title"] or ""
    local_path = organizer.save_text(seq_id, province, city, page_title,
                                     extracted["clean_text"], url_str, fetch_result.encoding)

    # Scan the fetched HTML for embedded PDF/DOC/DOCX/WPS links and download them
    file_links = extractor.find_file_links(fetch_result.text, url_str)
    downloaded_paths = []
    for flink in file_links:
        if _is_skip_domain(flink["url"]):
            continue
        limiter.wait()
        ext = os.path.splitext(urlparse(flink["url"]).path)[1] or ".pdf"
        fname = organizer.make_download_name(seq_id, province, city, flink["text"] or period, ext)
        dl_result = download_file(flink["url"], organizer.files_subdir(province, city), fname)
        if dl_result.status == "SUCCESS":
            limiter.report_success()
            downloaded_paths.append(dl_result.local_path)
        else:
            limiter.report_error()

    all_paths = [local_path] + downloaded_paths
    result["本地路径"] = " | ".join(all_paths)
    result["抓取状态"] = "SUCCESS"
    return result


def _run_pool(df, cp, checkpoint_path, limiter, workers: int) -> dict:
    """Process all rows using thread pool. Returns {链接序号: result_dict}."""
    all_results = {}
    rows_to_process = []

    for idx, row in df.iterrows():
        if not running:
            break
        seq_id = row["链接序号"]
        if checkpoint.is_done(cp, idx):
            continue
        rows_to_process.append((idx, row))

    total = len(rows_to_process)
    if total == 0:
        return all_results

    completed_count = 0

    def process_item(idx_row):
        nonlocal completed_count
        idx, row = idx_row
        seq_id = row["链接序号"]

        result = _process_one(row, limiter)

        with lock:
            completed_count += 1
            status = result["抓取状态"]
            city = row["城市"]
            period = row["所属时期"]
            local = result.get("本地路径", "")
            print(f"[{completed_count}/{total}] {city} - {period} -> {status} {local}")

            if status == "SUCCESS":
                checkpoint.mark(cp, idx, "fetched", local)
            elif status in ("SKIPPED", "SKIPPED_DOMAIN", "NO_URL"):
                checkpoint.mark(cp, idx, "skipped")
            else:
                checkpoint.mark(cp, idx, "failed")

            if completed_count % 20 == 0:
                checkpoint.save(cp, checkpoint_path)

        return seq_id, result

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(process_item, item): item for item in rows_to_process}
        for future in as_completed(futures):
            if not running:
                executor.shutdown(wait=False, cancel_futures=True)
                break
            try:
                seq_id, result = future.result()
                all_results[seq_id] = result
            except Exception as e:
                idx, row = futures[future]
                print(f"ERROR: {row['城市']} - {row['所属时期']}: {e}")

    checkpoint.save(cp, checkpoint_path)
    return all_results


def _write_back_to_excel(results: dict):
    """Write crawl results back to the original Excel's 备注 column."""
    input_path = str(config.DATA_FILE)
    if input_path.endswith(".xls"):
        # Read as xls, write as xlsx
        df_all = pd.read_excel(input_path)
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write header
        for col_idx, col_name in enumerate(df_all.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data rows with updated 备注
        for row_idx, (_, row) in enumerate(df_all.iterrows(), 2):
            for col_idx, col_name in enumerate(df_all.columns, 1):
                val = row[col_name]
                if pd.isna(val):
                    ws.cell(row=row_idx, column=col_idx, value="")
                else:
                    ws.cell(row=row_idx, column=col_idx, value=val)

            seq_id = row["链接序号"]
            if seq_id in results:
                r = results[seq_id]
                note = f"状态:{r['抓取状态']} | 路径:{r.get('本地路径','')}"
                # Find 备注 column index
                备注_col = list(df_all.columns).index("备注") + 1
                ws.cell(row=row_idx, column=备注_col, value=note)

        output_path = input_path.replace(".xls", "_已爬取.xlsx")
        wb.save(output_path)
        print(f"结果已写入: {output_path}")
        return output_path
    else:
        # Open xlsx directly
        df_all = pd.read_excel(input_path)
        # Create a map
        for _, row in df_all.iterrows():
            seq_id = row["链接序号"]
            if seq_id in results:
                r = results[seq_id]
                note = f"状态:{r['抓取状态']} | 路径:{r.get('本地路径','')}"
                df_all.at[_, "备注"] = note

        output_path = input_path.replace(".xlsx", "_已爬取.xlsx")
        df_all.to_excel(output_path, index=False)
        print(f"结果已写入: {output_path}")
        return output_path


def run(args) -> int:
    print(f"读取数据：{config.DATA_FILE}")
    df = pd.read_excel(config.DATA_FILE)

    if args.city:
        df = df[df["城市"] == args.city]
        if len(df) == 0:
            print(f"未找到城市：{args.city}")
            return 1
        print(f"过滤城市：{args.city}，共 {len(df)} 条")

    if args.limit:
        df = df.head(args.limit)

    total = len(df)
    workers = args.workers
    print(f"共 {total} 条，{workers} 个并发线程")
    print()

    cp = checkpoint.load(config.CHECKPOINT_FILE)
    if args.resume and cp.rows:
        done = sum(1 for r in cp.rows.values() if r.status in ("fetched", "skipped"))
        print(f"断点续传：已完成 {done}/{total}")

    # Thread-safe rate limiter shared across workers
    limiter = rate_limiter.RateLimiter(
        min_delay=0.5,
        max_delay=1.5,
    )

    start_time = datetime.now()

    results = _run_pool(df, cp, config.CHECKPOINT_FILE, limiter, workers)

    elapsed = (datetime.now() - start_time).total_seconds()

    fetched = sum(1 for r in results.values() if r["抓取状态"] == "SUCCESS")
    failed = sum(1 for r in results.values() if r["抓取状态"] == "FAILED")
    skipped = sum(1 for r in results.values() if r["抓取状态"] in ("NO_URL", "SKIPPED_DOMAIN", "SKIPPED"))
    aborted = sum(1 for r in results.values() if r["抓取状态"] == "ABORTED")

    print(f"\n{'='*50}")
    print(f"完成。耗时 {elapsed:.0f} 秒 ({elapsed/60:.1f} 分钟)")
    print(f"  成功: {fetched}")
    print(f"  失败: {failed}")
    print(f"  跳过: {skipped}")
    print(f"  中断: {aborted}")
    print(f"  合计: {len(results)}")
    print()

    if results:
        _write_back_to_excel(results)

    return 0


def main():
    parser = argparse.ArgumentParser(description="中国地级市五年规划文件爬虫")
    parser.add_argument("--city", type=str, default=None, help="只处理指定城市")
    parser.add_argument("--resume", action="store_true", help="从断点续传")
    parser.add_argument("--limit", type=int, default=None, help="限制处理条数")
    parser.add_argument("--workers", type=int, default=5, help="并发线程数 (默认5)")
    args = parser.parse_args()
    sys.exit(run(args))


if __name__ == "__main__":
    main()
